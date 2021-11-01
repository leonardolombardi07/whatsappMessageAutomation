import time #Permite pausar o programa por um certo número de segundos
from selenium import webdriver #Permite criar um bot que abre o google chrome
from selenium.webdriver.support.ui import WebDriverWait #Permite aguardar a presença de um elemento
from selenium.webdriver.support import expected_conditions as ec #Permite especificar as condições em que aguardaremos um elementoo
from selenium.webdriver.common.by import By #Permite especificar o tipo de elemento que queremos encontrar
from selenium.webdriver.common.keys import Keys #Permite pressionar enter, por exemplo
from selenium.webdriver.chrome.options import Options #Permite especificar a configuração do nosso driver
import openpyxl #Permite acessar uma planilha excel
from webdriver_manager.chrome import ChromeDriverManager

# Começar leitura na função Main (linha 116 ou algo parecido)

def openDriverAndScanQrCodeManually(seconds = 15):
    '''
    Essa função abre a página principal do whatsapp e espera "seconds" segundos
    pro usuário scannear manualmente o QR code com a camera do seu celular
    '''

    driver.get('https://web.whatsapp.com/') #Inicializa o driver entrando na página principal do WhatApp
    time.sleep(seconds) #Aguarda um intervalo de segundos (para que o QR code seja scanneado)

def getPhoneList():
    '''
    Essa função abre um arquivo excel contendo telefones
    na coluna A e retorna uma lista com esses telefones
    '''

    wb = openpyxl.load_workbook('ContatosPame.xlsx')
    phones_tab = wb['telefones']

    phone_list = []
    for row in range(2,phones_tab.max_row + 1):
        row_content = phones_tab['A' + str(row)].value
        phone_list.append(str(row_content))
        
    phone_list =(phone_list)
    return (phone_list)

def getNameList():
    '''
    Essa função abre um arquivo excel contendo os nomes dos detentores
    dos telefones na coluna B e retorna uma lista com esses nomes
    '''

    wb = openpyxl.load_workbook('contatosPame.xlsx')
    phones_tab = wb['telefones']

    name_list = []
    for row in range(2,phones_tab.max_row + 1):
        row_content = phones_tab['B' + str(row)].value
        name_list.append(str(row_content))
        
    name_list = list(name_list)
    return (name_list)

def loopPhoneListAndSendMessage():
    '''
    Essa função itera sobre uma lista com telefones (obtida utilizando a função getPhoneList) e envia,
    pelo whatsapp, uma mensagem específica pra cada telefone
    '''

    # Obtendo dados (telefones e nomes)
    phonesList = getPhoneList() #Lista com telefones
    namesList = getNameList() #Lista com nomes

    print(phonesList)

    # Iterando sobre todos os numeros
    for i in range(len(phonesList)):
        '''
        Para enviar uma mensagem com espaços entre linhas, você deverá encodifica-la.
        Utilize o site https://www.url-encode-decode.com/ para isso.
        
        Se você quiser partes da mensagem em negrito/itálico/etc, lembre-se de
        colocar um * para negrito ou _ para itálico antes do texto quando encodificar
        seu texto no site https://www.url-encode-decode.com/

        Assim como phonesList[i] representa um número de telefone, namesList[i]
        representa o nome do detentor desse telefone. Você pode utilizar esse nome
        dentro da mensagem concatenando-o à string obtida em https://www.url-encode-decode.com/

        EXEMPLO:
            MENSAGEM: OLA, PESSOA
                      COMO VOCÊ ESTÁ?
            (PESSOA nesse caso é um nome genérico)

            MENSAGEM ENCODIFICADA: "OLA+PESSOA%0D%0ACOMO+VOC%C3%8A+EST%C3%81%3F"

            UTILIZANDO NOME DE USUÁRIO NA MENSAGEM: "OLA+" namesList[i] + "%0D%0ACOMO+VOC%C3%8A+EST%C3%81%3F"
            isto é, a variável que contêm a mensagem deverá ser:
                message = "OLA+" namesList[i] + "%0D%0ACOMO+VOC%C3%8A+EST%C3%81%3F"
        '''
        
        message = "Ola, Leonardo" # Mensagem que será enviada pra cada telefone
        try:
            # 1) Abre a página do whatsapp com o número e a mensagem especifacados
            driver.get('https://web.whatsapp.com/send?phone=' + str(phonesList[i]) + '&text=' + message)

            # 2) Aguarda no máximo 15 segundos até encontrar caixa de texto

            '''
            MUITO IMPORTANTE: na penultima string da variável "xpath" abaixo, um trecho da mensagem
            enviada (não encodificado) deve ser colocado. Só assim o programa consegue reconhecer
            onde está o textbox com a mensagem enviada e então apertar enter
            '''
            
            xpath = "//*[contains(text(), " + "'"+ "Ola" + "')]"
            wait = WebDriverWait(driver, 15) #Tenta encontrar a caixa de texto. Se passar de 15 segundos, vai pra próxima mensagem
            whatsapp_message_box = wait.until(ec.visibility_of_any_elements_located((By.XPATH, xpath)))
            whatsapp_message_box = whatsapp_message_box[-1] #Seleciona a caixa de texto

            # 3) Envia a mensagem
            whatsapp_message_box.send_keys(Keys.ENTER)

        except:
            next


def main():
    '''
    Essa é a função principal, executada sempre que esse arquivo é executado no python
    '''
    
    # driver é uma variável que vai permitir que o Selenium abra o browser (chrome, nesse caso)
    global driver 


    # chrome_options permite especificar um determinado comportamento pro driver
    chrome_options = Options() 
    chrome_options.set_capability('unhandledPromptBehavior', 'accept') 
    
    # definimos o driver pelo selenium passando as opções especifcadas
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    # driver = webdriver.Chrome(options=chrome_options)

    # funções que são rodadas em ordem
    openDriverAndScanQrCodeManually()
    loopPhoneListAndSendMessage()

if __name__ == '__main__':
    main()
